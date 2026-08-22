from flask import Flask, render_template, request, redirect, url_for, flash, send_file
from pymongo import MongoClient
import certifi
from datetime import datetime
import pytz
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from werkzeug.utils import secure_filename
import io
import re
import os

app = Flask(__name__)
app.secret_key = "secret_key_123"

# File Upload Configuration
UPLOAD_FOLDER = 'static/uploads'
ALLOWED_EXTENSIONS = {'pdf', 'docx', 'doc', 'png', 'jpg', 'jpeg', 'xlsx'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# Ensure upload directory exists
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# MongoDB Connection String
MONGO_URI = os.getenv("MONGO_URI", "mongodb+srv://pratheeshh583225106064_db_user:plcKiS5p7c4S0G15@bitron.gge3k34.mongodb.net/?appName=Bitron")

client = MongoClient(
    MONGO_URI,
    tlsCAFile=certifi.where()
)

db = client['attendance_system']
members_col = db['members']
attendance_col = db['attendance']
system_col = db['system_config']
reports_col = db['daily_reports']

IST = pytz.timezone('Asia/Kolkata')

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def check_and_reset_weekly():
    now = datetime.now(IST)
    current_year, current_week, _ = now.isocalendar()
    
    config = system_col.find_one({'type': 'weekly_tracker'})
    
    if not config:
        system_col.insert_one({'type': 'weekly_tracker', 'year': current_year, 'week': current_week})
    else:
        if config['year'] < current_year or config['week'] < current_week:
            attendance_col.delete_many({})
            reports_col.delete_many({})
            system_col.update_one(
                {'type': 'weekly_tracker'},
                {'$set': {'year': current_year, 'week': current_week}}
            )

def init_db():
    default_members = [
        {'name': 'Sivamani V', 'email': 'sivamani1234@gmail.com', 'role': 'CEO & Founder'},
        {'name': 'Anusha', 'email': 'anusha1234@gmail.com', 'role': 'Technical coordinator'},
        {'name': 'Dharsana G', 'email': 'dharsana1234@gmail.com', 'role': 'IoT Engineer'},
        {'name': 'Aisha mariyam', 'email': 'aisha1234@gmail.com', 'role': 'IoT Engineer'},
        {'name': 'Meenakshi Priyadarshini', 'email': 'meenakshi1234@gmail.com', 'role': 'PCB Designer'},
        {'name': 'Shyam kumar M', 'email': 'shyam1234@gmail.com', 'role': 'PCB designer'},
        {'name': 'Pratheesh H', 'email': 'pratheesh1234@gmail.com', 'role': 'PCB designer'}
    ]
    
    for member in default_members:
        members_col.update_one(
            {'email': member['email']},
            {'$setOnInsert': member},
            upsert=True
        )

init_db()

@app.route('/')
def home():
    return render_template('bitron-website-1.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        name = request.form.get('name')
        email = request.form.get('email')
        
        if not name or not email:
            flash("Please enter both Name and Email!")
            return redirect(url_for('login'))

        email_pattern = r'^[a-zA-Z]+[0-9]+@gmail\.com$'
        if not re.match(email_pattern, email):
            flash("❌ Invalid Email Format! Email must be like: name1234@gmail.com")
            return redirect(url_for('login'))

        user = members_col.find_one({'email': email})
        
        if not user:
            flash("❌ Access Denied! You are not an authorized member.")
            return redirect(url_for('login'))
        else:
            members_col.update_one({'email': email}, {'$set': {'name': name}})
            return redirect(url_for('dashboard', email=email))
        
    return render_template('login.html')

@app.route('/dashboard', methods=['GET', 'POST'])
def dashboard():
    email = request.args.get('email') or request.form.get('email')
    if not email:
        return redirect(url_for('login'))
        
    user = members_col.find_one({'email': email})
    if not user:
        return redirect(url_for('login'))
        
    history_cursor = attendance_col.find({'email': email}).sort('_id', -1)
    history = [[doc['date'], doc['time'], doc['status']] for doc in history_cursor]
    
    now = datetime.now(IST)
    today_date = now.strftime("%Y-%m-%d")
    
    today_rec = attendance_col.find_one({'email': email, 'date': today_date})
    today_attendance = (today_rec['status'], today_rec['time']) if today_rec else None
    
    today_report_rec = reports_col.find_one({'email': email, 'date': today_date})
    today_work_report = today_report_rec['work_done'] if today_report_rec else None
    today_file_path = today_report_rec.get('file_path') if today_report_rec else None

    selected_date = request.form.get('selected_date', today_date)
    
    date_rec = attendance_col.find_one({'email': email, 'date': selected_date})
    date_record = (date_rec['status'], date_rec['time']) if date_rec else None
    
    total_present = sum(1 for row in history if row[2] == 'Present')
    total_absent = sum(1 for row in history if row[2] == 'Absent')
    
    user_tuple = (user['name'], user['email'], user['role'])
    
    return render_template('dashboard.html', user=user_tuple, history=history, 
                           today_date=today_date, selected_date=selected_date,
                           date_record=date_record, today_attendance=today_attendance,
                           today_work_report=today_work_report, today_file_path=today_file_path,
                           total_present=total_present, total_absent=total_absent)

@app.route('/mark_attendance', methods=['POST'])
def mark_attendance():
    check_and_reset_weekly()

    email = request.form.get('email')
    status = request.form.get('status')
    
    now = datetime.now(IST)
    date = now.strftime("%Y-%m-%d")
    time = now.strftime("%I:%M:%S %p")
    
    existing_record = attendance_col.find_one({'email': email, 'date': date})
    
    if existing_record:
        flash("❌ Attendance already marked for today!")
    else:
        attendance_col.insert_one({
            'email': email,
            'date': date,
            'time': time,
            'status': status
        })
        flash("✅ Attendance Marked Successfully!")
        
    return redirect(url_for('dashboard', email=email))

@app.route('/submit_daily_report', methods=['POST'])
def submit_daily_report():
    check_and_reset_weekly()

    email = request.form.get('email')
    work_done = request.form.get('work_done')
    file = request.files.get('report_file')
    
    file_path = None
    if file and file.filename != '' and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        unique_filename = f"{datetime.now(IST).strftime('%Y%m%d%H%M%S')}_{filename}"
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], unique_filename)
        file.save(file_path)

    now = datetime.now(IST)
    date = now.strftime("%Y-%m-%d")
    time = now.strftime("%I:%M:%S %p")

    reports_col.update_one(
        {'email': email, 'date': date},
        {'$set': {
            'email': email, 
            'date': date, 
            'time': time, 
            'work_done': work_done,
            'file_path': file_path
        }},
        upsert=True
    )
    
    flash("✅ Daily Work Report Submitted Successfully!")
    return redirect(url_for('dashboard', email=email))

@app.route('/report')
def report():
    reports = []
    attendance_records = attendance_col.find().sort('_id', -1)
    
    for att in attendance_records:
        member = members_col.find_one({'email': att['email']})
        name = member['name'] if member else 'Unknown'
        role = member['role'] if member else 'Unknown'
        
        rep = reports_col.find_one({'email': att['email'], 'date': att['date']})
        work_done = rep['work_done'] if rep else 'Not Submitted'
        file_path = rep.get('file_path') if rep else None
        
        reports.append((name, role, att['email'], att['date'], att['time'], att['status'], work_done, file_path))
        
    return render_template('report.html', reports=reports)

@app.route('/download_excel')
def download_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance & Work Report"
    
    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )
    
    headers = ['Name', 'Role', 'Email', 'Date', 'Time (IST)', 'Status', 'Daily Work Report', 'Attachment File']
    ws.append(headers)
    
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    attendance_records = attendance_col.find().sort('_id', -1)
    data_font = Font(name="Arial", size=10)
    
    for row_idx, att in enumerate(attendance_records, start=2):
        member = members_col.find_one({'email': att['email']})
        name = member['name'] if member else 'Unknown'
        role = member['role'] if member else 'Unknown'
        
        rep = reports_col.find_one({'email': att['email'], 'date': att['date']})
        work_done = rep['work_done'] if rep else 'Not Submitted'
        file_path = rep.get('file_path', 'No File') if rep else 'No File'
        
        row_data = [name, role, att['email'], att['date'], att['time'], att['status'], work_done, file_path]
        ws.append(row_data)
        
        row_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid") if row_idx % 2 == 0 else PatternFill(fill_type=None)
        
        for cell in ws[row_idx]:
            cell.font = data_font
            cell.border = thin_border
            cell.fill = row_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 6, 16)
        
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"Attendance_Work_Report_{datetime.now(IST).strftime('%Y-%m-%d')}.xlsx"
    
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

if __name__ == '__main__':
    app.run(debug=True)
        
